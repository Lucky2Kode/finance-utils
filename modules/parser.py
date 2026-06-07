"""
Parse plain-text or Excel stock watchlists into structured Company records.

Supported text line formats:
  Company Name (EXCHANGE: TICKER)   e.g. "Bloom Energy (NYSE: BE)"
  Company Name (TICKER)             e.g. "Palantir (PLTR)"
  (EXCHANGE: TICKER)                e.g. "(NYSEARCA: XLE)"
  EXCHANGE: TICKER                  e.g. "NASDAQ: MRVL"

Supported Excel input columns (case-insensitive, any order):
  Symbol / Ticker, Name, Exchange
"""

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

# Format 1 & 2: optional name before (…)
_WITH_PARENS_RE = re.compile(r"^(.*?)\s*\(([^)]+)\)\s*$")
# Format 3: bare EXCHANGE: TICKER with no parens and no name
_BARE_RE = re.compile(r"^([A-Z]+):\s*([A-Z0-9.]+)\s*$")

_INC_RE = re.compile(r"\bInc\.\s*")
_CORP_RE = re.compile(r"\bCorp\b\s*")


@dataclass(frozen=True)
class Company:
    ticker: str
    name: str
    exchange: str

    def __str__(self) -> str:
        return f"{self.ticker} | {self.name} | {self.exchange}"


def _clean_name(raw: str) -> str:
    name = _INC_RE.sub("", raw)
    name = _CORP_RE.sub("", name)
    return name.strip()


def _parse_line(line: str) -> Optional[Company]:
    line = line.strip()

    # Format 1 & 2: anything with parentheses — "Name (EXCHANGE: TICKER)" or "(EXCHANGE: TICKER)"
    m = _WITH_PARENS_RE.match(line)
    if m:
        raw_name = m.group(1).strip()
        paren_content = m.group(2).strip()
        if ":" in paren_content:
            exchange, ticker = paren_content.split(":", 1)
            exchange, ticker = exchange.strip(), ticker.strip()
        else:
            exchange, ticker = "", paren_content
        return Company(ticker=ticker, name=_clean_name(raw_name), exchange=exchange)

    # Format 3: bare "EXCHANGE: TICKER" with no parens
    m = _BARE_RE.match(line)
    if m:
        exchange, ticker = m.group(1).strip(), m.group(2).strip()
        return Company(ticker=ticker, name="", exchange=exchange)

    return None


def read_and_parse(file_path: str | Path) -> tuple[list[Company], list[str]]:
    """
    Read a plain-text watchlist file and return:
      - a deduplicated, ticker-sorted list of Company objects
      - a list of raw lines that could not be parsed

    Deduplication is by (ticker, exchange) — keeps first occurrence.
    """
    path = Path(file_path)
    lines = [l for l in path.read_text(encoding="utf-8").splitlines() if l.strip()]

    seen: set[tuple[str, str]] = set()
    companies: list[Company] = []
    unparsed: list[str] = []

    for line in lines:
        company = _parse_line(line)
        if company is None:
            unparsed.append(line.strip())
            continue
        key = (company.ticker.upper(), company.exchange.upper())
        if key in seen:
            continue
        seen.add(key)
        companies.append(company)

    companies.sort(key=lambda c: c.ticker.upper())
    return companies, unparsed


def read_excel_and_parse(file_path: str | Path) -> tuple[list[Company], list[tuple]]:
    """
    Read Company records from an Excel input file.

    Sheet priority: Companies → Stocks → first sheet.
    Columns matched case-insensitively: Symbol/Ticker, Name, Exchange.

    Returns:
        companies   : deduplicated, ticker-sorted list of Company objects
        failed_rows : raw row tuples with no valid Symbol (kept in input file)
    """
    from openpyxl import load_workbook

    path = Path(file_path)
    wb = load_workbook(path)

    sheet_name = next(
        (s for s in ["Companies", "Stocks"] if s in wb.sheetnames),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]

    # Map header names to column indices (0-based)
    raw_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h or "").strip().lower() for h in raw_headers]

    col_ticker = next((i for i, h in enumerate(headers) if h in ("symbol", "ticker")), None)
    col_name = next((i for i, h in enumerate(headers) if h == "name"), None)
    col_exchange = next((i for i, h in enumerate(headers) if h == "exchange"), None)

    seen: set[tuple[str, str]] = set()
    companies: list[Company] = []
    failed_rows: list[tuple] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker = str(row[col_ticker] or "").strip() if col_ticker is not None else ""
        if not ticker:
            failed_rows.append(row)
            continue

        name = str(row[col_name] or "").strip() if col_name is not None else ""
        exchange = str(row[col_exchange] or "").strip() if col_exchange is not None else ""

        key = (ticker.upper(), exchange.upper())
        if key in seen:
            continue
        seen.add(key)
        companies.append(Company(ticker=ticker, name=name, exchange=exchange))

    wb.close()
    companies.sort(key=lambda c: c.ticker.upper())
    return companies, failed_rows
