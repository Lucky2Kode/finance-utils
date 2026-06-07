"""
Convert a stock watchlist (plain-text or Excel) to a styled Excel spreadsheet.

Output file: A2ZStocks.xlsx (fixed, in outputs/)
Sheets:
  Companies  — Symbol | Name | Exchange | Date Added
               - Sorted by ticker
               - Dark header, alternating blue/white rows
               - Incremental: re-runs only append symbols not already present
  Summary    — Letter | Count  (number of stocks per first letter) + Total row

Input types supported:
  .txt   — plain-text watchlist; parsed lines cleared, unparsed lines kept
  .xlsx  — Excel sheet (Companies/Stocks); processed rows removed from input file
"""

from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from modules.parser import read_and_parse, read_excel_and_parse, Company

_OUTPUT_FILENAME = "A2ZStocks.xlsx"
_EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}

_HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_EVEN_FILL = PatternFill("solid", fgColor="DCE6F1")
_ODD_FILL = PatternFill("solid", fgColor="FFFFFF")
_COMPANIES_HEADERS = ["Symbol", "Name", "Exchange", "Date Added"]
_SUMMARY_HEADERS = ["Letter", "Count"]


# ---------------------------------------------------------------------------
# Excel writing helpers
# ---------------------------------------------------------------------------

def _style_header(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _auto_size(ws, num_cols: int) -> None:
    for col_idx in range(1, num_cols + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3


def _load_existing(output_path: Path) -> dict[str, dict]:
    """Read the Companies sheet and return {TICKER: {name, exchange, date_added}}."""
    existing: dict[str, dict] = {}
    if not output_path.exists():
        return existing

    wb = load_workbook(output_path)
    if "Companies" not in wb.sheetnames:
        return existing

    ws = wb["Companies"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    def col(name):
        return headers.index(name)

    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker = row[col("Symbol")]
        if ticker:
            existing[ticker.upper()] = {
                "name": row[col("Name")],
                "exchange": row[col("Exchange")],
                "date_added": row[col("Date Added")],
            }
    return existing


def _write_companies(ws, records: list[tuple[str, dict]]) -> None:
    _style_header(ws, _COMPANIES_HEADERS)
    for row_idx, (ticker, data) in enumerate(records, start=2):
        fill = _EVEN_FILL if row_idx % 2 == 0 else _ODD_FILL
        for col_idx, value in enumerate(
            [ticker, data["name"], data["exchange"], data["date_added"]], start=1
        ):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
    _auto_size(ws, len(_COMPANIES_HEADERS))


def _write_summary(ws, records: list[tuple[str, dict]]) -> None:
    _style_header(ws, _SUMMARY_HEADERS)
    letter_counts: dict[str, int] = {}
    for ticker, _ in records:
        letter = ticker[0].upper()
        letter_counts[letter] = letter_counts.get(letter, 0) + 1

    for row_idx, letter in enumerate(sorted(letter_counts), start=2):
        fill = _EVEN_FILL if row_idx % 2 == 0 else _ODD_FILL
        ws.cell(row=row_idx, column=1, value=letter).fill = fill
        ws.cell(row=row_idx, column=2, value=letter_counts[letter]).fill = fill

    # Total row
    total_row = ws.max_row + 1
    for col, value in enumerate(["Total", sum(letter_counts.values())], start=1):
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    _auto_size(ws, len(_SUMMARY_HEADERS))


def _save_output(output_path: Path, existing: dict[str, dict]) -> None:
    records = sorted(existing.items(), key=lambda x: x[0])
    wb = Workbook()
    ws_companies = wb.active
    ws_companies.title = "Companies"
    _write_companies(ws_companies, records)
    ws_summary = wb.create_sheet("Summary")
    _write_summary(ws_summary, records)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Input file cleanup helpers
# ---------------------------------------------------------------------------

def _clear_txt_input(input_path: Path, unparsed: list[str], unparsed_path: Path) -> None:
    """Rewrite txt input keeping only unparsed lines; update Unparsed.txt."""
    if unparsed:
        unparsed_path.write_text("\n".join(unparsed) + "\n", encoding="utf-8")
    elif unparsed_path.exists():
        unparsed_path.unlink()

    input_path.write_text(
        "\n".join(unparsed) + ("\n" if unparsed else ""), encoding="utf-8"
    )


def _clear_excel_input(input_path: Path, failed_rows: list[tuple]) -> None:
    """Rewrite input Excel keeping only the header + rows that had no valid Symbol."""
    wb = load_workbook(input_path)
    sheet_name = next(
        (s for s in ["Companies", "Stocks"] if s in wb.sheetnames),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]

    # Delete all data rows then re-insert failed rows
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_idx, row_data in enumerate(failed_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(input_path)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_watchlist_to_excel(
    input_path: str | Path,
    output_path: Optional[str | Path] = None,
) -> Path:
    """
    Read *input_path* (.txt or .xlsx) and incrementally update A2ZStocks.xlsx.

    .txt input  — parses text lines; parsed lines cleared, unparsed lines kept.
    .xlsx input — reads Symbol/Name/Exchange rows; processed rows removed from file.

    Stocks already present (matched by Symbol) are skipped.
    New stocks are appended with today's date.
    Summary sheet is rebuilt on every run.
    """
    input_path = Path(input_path)

    if output_path is None:
        output_path = Path(__file__).parent.parent / "outputs" / _OUTPUT_FILENAME
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    existing = _load_existing(output_path)
    before = len(existing)
    today = date.today().isoformat()
    is_excel = input_path.suffix.lower() in _EXCEL_EXTENSIONS

    # --- Parse input ---
    if is_excel:
        new_companies, failed_rows = read_excel_and_parse(input_path)
        skipped_label = f"failed rows  : {len(failed_rows)} (kept in input file)"
    else:
        new_companies, unparsed = read_and_parse(input_path)
        skipped_label = f"Unparsed     : {len(unparsed)} lines" + (
            f" → {output_path.parent / 'Unparsed.txt'}" if unparsed else " (none)"
        )

    # --- Merge ---
    for company in new_companies:
        key = company.ticker.upper()
        if key not in existing:
            existing[key] = {
                "name": company.name,
                "exchange": company.exchange,
                "date_added": today,
            }

    added = len(existing) - before

    # --- Clean up input file ---
    if is_excel:
        _clear_excel_input(input_path, failed_rows)
        cleared_msg = f"Cleared      : processed rows removed from {input_path.name}"
    else:
        _clear_txt_input(input_path, unparsed, output_path.parent / "Unparsed.txt")
        cleared_msg = f"Cleared      : parsed lines removed from {input_path.name}"

    # --- Write output ---
    _save_output(output_path, existing)

    print(f"Input file   : {input_path} ({'Excel' if is_excel else 'text'})")
    print(f"Existing     : {before} stocks")
    print(f"New added    : {added} stocks")
    print(f"Total unique : {len(existing)} stocks")
    print(skipped_label)
    print(cleared_msg)
    print(f"Saved        : {output_path}")
    return output_path


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python -m modules.watchlist_to_excel <input-file> [output-file]")
        sys.exit(1)
    export_watchlist_to_excel(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
